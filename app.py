import os
import re
import difflib
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from fastapi import FastAPI, HTTPException, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List, Optional
import generate_excel

app = FastAPI(title="AADHAN FIRE WORKS - Order Dashboard")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PRIMARY_EXCEL_PATH = os.path.join(BASE_DIR, "Firecrackers_Catalog_and_Orders.xlsx")
UPDATED_EXCEL_PATH = os.path.join(BASE_DIR, "Firecrackers_Catalog_and_Orders_Updated.xlsx")
STATIC_IMG_DIR = os.path.join(BASE_DIR, "static", "images")

# Admin Owner Credentials
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "aadhan@2026")

def get_active_excel_path():
    if os.path.exists(UPDATED_EXCEL_PATH):
        return UPDATED_EXCEL_PATH
    return PRIMARY_EXCEL_PATH

# Ensure Excel file exists on startup
if not os.path.exists(PRIMARY_EXCEL_PATH) and not os.path.exists(UPDATED_EXCEL_PATH):
    generate_excel.extract_and_generate()

class OrderItem(BaseModel):
    category: str
    cracker_name: str
    unit_price: float
    quantity: int

class OrderCreateRequest(BaseModel):
    buyer_name: str
    contact_number: str
    category: Optional[str] = None
    cracker_name: Optional[str] = None
    unit_price: Optional[float] = None
    quantity: Optional[int] = None
    items: Optional[List[OrderItem]] = None

class AdminAuthRequest(BaseModel):
    username: str
    password: str

def normalize_image_key(s: str) -> str:
    s = str(s).lower()
    s = s.replace('1/2', '12').replace('3/2', '32').replace('1 1/2', '112').replace('3 1/2', '312')
    return re.sub(r'[^a-z0-9]', '', s)

def get_image_url(category_name):
    if not os.path.exists(STATIC_IMG_DIR):
        return "/static/images/7_cm_sparklers.png"

    cat_clean = str(category_name).strip()
    cat_norm = normalize_image_key(cat_clean)
    
    files = os.listdir(STATIC_IMG_DIR)
    img_map = {}
    for f in files:
        base = os.path.splitext(f)[0]
        norm = normalize_image_key(base)
        img_map[norm] = f

    # 1. Exact normalized match
    if cat_norm in img_map:
        return f"/static/images/{img_map[cat_norm]}"

    # 2. Substring match where norm is in cat_norm or cat_norm is in norm
    for norm, filename in img_map.items():
        if len(norm) > 4 and (norm in cat_norm or cat_norm in norm):
            return f"/static/images/{filename}"

    # 3. Fuzzy similarity matching with numeric constraint
    num_cat = set(re.findall(r'\d+', cat_clean))
    best_match = None
    best_ratio = 0.0

    for norm, filename in img_map.items():
        num_img = set(re.findall(r'\d+', norm))
        if num_cat or num_img:
            if num_cat != num_img:
                continue

        ratio = difflib.SequenceMatcher(None, cat_norm, norm).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = filename

    if best_match and best_ratio > 0.4:
        return f"/static/images/{best_match}"

    return "/static/images/7_cm_sparklers.png"

@app.get("/api/catalog")
def get_catalog():
    excel_path = get_active_excel_path()
    if not os.path.exists(excel_path):
        generate_excel.extract_and_generate()
        excel_path = get_active_excel_path()
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    if "Product Catalog" not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=500, detail="Product Catalog sheet missing")
    
    ws = wb["Product Catalog"]
    products = []
    categories_set = set()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        box_photo, item_id, factory, category, name, rate, per, case_content = row[:8]
        cat_clean = generate_excel.normalize_category(str(category))
        categories_set.add(cat_clean)
        products.append({
            "item_id": str(item_id),
            "factory": str(factory),
            "category": cat_clean,
            "name": str(name).strip(),
            "price": float(rate or 0.0),
            "per": str(per or "").strip(),
            "case_content": str(case_content or "").strip(),
            "image_url": get_image_url(cat_clean)
        })
    wb.close()
        
    return {
        "categories": sorted(list(categories_set)),
        "products": products
    }

@app.post("/api/admin/verify")
def verify_admin(req: AdminAuthRequest):
    if req.username == ADMIN_USERNAME and req.password == ADMIN_PASSWORD:
        return {"success": True, "token": ADMIN_PASSWORD}
    raise HTTPException(status_code=401, detail="Invalid Owner Credentials!")

@app.post("/api/orders")
def place_order(req: OrderCreateRequest):
    if not req.buyer_name or not req.contact_number:
        raise HTTPException(status_code=400, detail="Buyer Name and Contact Number are required.")
    
    order_items = []
    if req.items and len(req.items) > 0:
        order_items = req.items
    elif req.category and req.cracker_name and req.unit_price is not None and req.quantity is not None:
        order_items.append(OrderItem(
            category=req.category,
            cracker_name=req.cracker_name,
            unit_price=req.unit_price,
            quantity=req.quantity
        ))
    else:
        raise HTTPException(status_code=400, detail="Please select a cracker and enter quantity.")

    excel_path = get_active_excel_path()
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception:
        excel_path = PRIMARY_EXCEL_PATH
        wb = openpyxl.load_workbook(excel_path)

    if "Order Details" not in wb.sheetnames:
        ws_orders = wb.create_sheet(title="Order Details")
        order_headers = ["Order ID", "Order Date & Time", "Buyer Name", "Contact Number", "Category", "Cracker Name", "Unit Price (₹)", "Quantity", "Total Amount (₹)"]
        ws_orders.append(order_headers)
    else:
        ws_orders = wb["Order Details"]

    existing_rows = max(0, ws_orders.max_row - 1)
    timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    created_orders = []
    grand_total = 0.0
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for i, item in enumerate(order_items):
        existing_rows += 1
        order_id = f"ORD-{existing_rows:04d}"
        item_total = round(item.unit_price * item.quantity, 2)
        grand_total += item_total
        
        row_data = [
            order_id,
            timestamp_str,
            req.buyer_name.strip(),
            req.contact_number.strip(),
            item.category.strip(),
            item.cracker_name.strip(),
            item.unit_price,
            item.quantity,
            item_total
        ]
        
        ws_orders.append(row_data)
        r_idx = ws_orders.max_row
        
        ws_orders.cell(row=r_idx, column=7).number_format = '₹#,##0.00'
        ws_orders.cell(row=r_idx, column=9).number_format = '₹#,##0.00'
        for c in range(1, 10):
            ws_orders.cell(row=r_idx, column=c).border = thin_border
            
        created_orders.append({
            "order_id": order_id,
            "date_time": timestamp_str,
            "buyer_name": req.buyer_name,
            "contact_number": req.contact_number,
            "category": item.category,
            "cracker_name": item.cracker_name,
            "unit_price": item.unit_price,
            "quantity": item.quantity,
            "total_amount": item_total,
            "image_url": get_image_url(item.category)
        })

    try:
        wb.save(excel_path)
    except PermissionError:
        excel_path = UPDATED_EXCEL_PATH
        wb.save(excel_path)
    wb.close()

    return {
        "success": True,
        "message": f"Successfully placed {len(created_orders)} item order(s)!",
        "orders": created_orders,
        "grand_total": grand_total
    }

@app.get("/api/download-excel")
def download_excel(key: Optional[str] = Query(None)):
    if key != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized. Owner login required.")

    excel_path = get_active_excel_path()
    if not os.path.exists(excel_path):
        generate_excel.extract_and_generate()
        excel_path = get_active_excel_path()
    return FileResponse(
        path=excel_path,
        filename="Firecrackers_Catalog_and_Orders.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
app.mount("/", StaticFiles(directory=BASE_DIR, html=True), name="static_root")

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
