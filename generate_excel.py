import os
import re
import difflib
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

pdf1_path = os.path.join(BASE_DIR, "SRI NARAYANA SPARKLERS FACTORY PRICELIST (1).pdf")
pdf2_path = os.path.join(BASE_DIR, "SRI NARAYANA SPARKLERS FACTORY PRICELIST (2).pdf")
excel_path = os.path.join(BASE_DIR, "Firecrackers_Catalog_and_Orders.xlsx")
STATIC_IMG_DIR = os.path.join(BASE_DIR, "static", "images")

def normalize_image_key(s: str) -> str:
    s = str(s).lower()
    s = s.replace('1/2', '12').replace('3/2', '32').replace('1 1/2', '112').replace('3 1/2', '312')
    return re.sub(r'[^a-z0-9]', '', s)

def find_category_image_path(category_name, static_dir=STATIC_IMG_DIR):
    if not os.path.exists(static_dir):
        return None

    cat_clean = str(category_name).strip()
    cat_norm = normalize_image_key(cat_clean)
    
    files = os.listdir(static_dir)
    img_map = {}
    for f in files:
        base = os.path.splitext(f)[0]
        norm = normalize_image_key(base)
        img_map[norm] = f

    if cat_norm in img_map:
        return os.path.join(static_dir, img_map[cat_norm])

    for norm, filename in img_map.items():
        if len(norm) > 4 and (norm in cat_norm or cat_norm in norm):
            return os.path.join(static_dir, filename)

    num_cat = set(re.findall(r'\d+', cat_clean))
    best_match = None
    best_ratio = 0.0

    for norm, filename in img_map.items():
        num_img = set(re.findall(r'\d+', norm))
        if num_cat or num_img:
            if num_cat != num_img:
                continue

        ratio = difflib.SequenceMatcher(None, cat_norm, norm).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = filename

    if best_match and best_ratio > 0.4:
        return os.path.join(static_dir, best_match)

    fallback = os.path.join(static_dir, "7_cm_sparklers.png")
    return fallback if os.path.exists(fallback) else None

def normalize_category(cat_name):
    c = cat_name.strip()
    if c.lower() in ["31/2\"pipe", "3 1/2\"pipe", "3 1/2 pipe", "31/2 pipe"]:
        return '3 1/2" Pipe'
    if c.lower() in ["11/2\"pipe", "1 1/2\"pipe", "1 1/2 pipe", "11/2 pipe"]:
        return '1 1/2" Pipe'
    return c.title() if not any(x in c for x in ['"', '1/2', 'Cm', 'Pcs']) else c

def extract_and_generate(target_path=excel_path):
    items = []

    # --- PARSE PDF 1 ---
    if os.path.exists(pdf1_path):
        with pdfplumber.open(pdf1_path) as pdf:
            current_category = "Sparklers"
            for page in pdf.pages:
                tables = page.extract_tables()
                for t in tables:
                    for row in t:
                        if not row or all(c is None for c in row):
                            continue
                        if row[0] == 'S.NO' or 'PARTICULARS' in str(row[1]):
                            continue
                        if row[0] and row[1] is None and row[2] is None:
                            current_category = normalize_category(row[0])
                            continue
                        if len(row) >= 5 and row[0] and row[0].isdigit():
                            name = row[1].strip() if row[1] else ""
                            rate = float(row[2].strip()) if row[2] and row[2].strip().isdigit() else (float(row[2]) if row[2] else 0.0)
                            per = row[3].strip() if row[3] else ""
                            case_content = row[4].strip() if row[4] else ""
                            items.append({
                                "factory": "AADHAN FIRE WORKS",
                                "category": normalize_category(current_category),
                                "name": name,
                                "rate": rate,
                                "per": per,
                                "case_content": case_content
                            })

    # --- PARSE PDF 2 ---
    if os.path.exists(pdf2_path):
        with pdfplumber.open(pdf2_path) as pdf:
            current_category = "Single Crackers"
            for page in pdf.pages:
                tables = page.extract_tables()
                for t in tables:
                    for row in t:
                        if not row or all(c is None for c in row):
                            continue
                        if row[0] == 'products' or 'price' in str(row[1]):
                            continue
                        if row[0] and (row[1] is None or str(row[1]).strip() == '') and row[2] is None:
                            cat_name = row[0].strip()
                            if cat_name.lower() not in ['1 | p age', '2 | p age', '3 | p age', '4 | p age']:
                                current_category = normalize_category(cat_name)
                            continue
                        if len(row) >= 4 and row[1] and str(row[1]).replace('.', '', 1).isdigit():
                            name = row[0].strip() if row[0] else ""
                            if name.lower() in ['sri narayana fireworks industries', 'aadhan fire works', '1 | p age', '2 | p age', '3 | p age', '4 | p age']:
                                continue
                            price_val = float(row[1].strip())
                            per = row[2].strip() if row[2] else ""
                            case_content = row[3].strip() if row[3] else ""
                            items.append({
                                "factory": "AADHAN FIRE WORKS",
                                "category": normalize_category(current_category),
                                "name": name,
                                "rate": price_val,
                                "per": per,
                                "case_content": case_content
                            })

    # Create Excel Workbook
    wb = openpyxl.Workbook()

    # Sheet 1: Product Catalog
    ws_catalog = wb.active
    ws_catalog.title = "Product Catalog"

    # Sheet 2: Order Details
    ws_orders = wb.create_sheet(title="Order Details")

    # Styles
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cat_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    order_header_fill = PatternFill(start_color="274E13", end_color="274E13", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Catalog Headers
    catalog_headers = ["Box Photo", "Item ID", "Factory / Unit", "Category", "Cracker Name", "Price / Rate (₹)", "Per Unit", "Case Content"]
    ws_catalog.append(catalog_headers)
    ws_catalog.row_dimensions[1].height = 24
    for col_num in range(1, len(catalog_headers) + 1):
        cell = ws_catalog.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = cat_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, item in enumerate(items, 1):
        row_idx = idx + 1
        ws_catalog.append([
            "", # Box photo column
            f"CRK-{idx:03d}",
            item["factory"],
            item["category"],
            item["name"],
            item["rate"],
            item["per"],
            item["case_content"]
        ])
        ws_catalog.cell(row=row_idx, column=6).number_format = '₹#,##0.00'
        ws_catalog.row_dimensions[row_idx].height = 42

        for c in range(1, 9):
            cell = ws_catalog.cell(row=row_idx, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        # Embed Carton Box image
        img_path = find_category_image_path(item["category"])

        if img_path and os.path.exists(img_path):
            try:
                xl_img = OpenPyXLImage(img_path)
                xl_img.width = 45
                xl_img.height = 45
                ws_catalog.add_image(xl_img, f"A{row_idx}")
            except Exception as e:
                pass

    # Order Details Headers
    order_headers = ["Order ID", "Order Date & Time", "Buyer Name", "Contact Number", "Category", "Cracker Name", "Unit Price (₹)", "Quantity", "Total Amount (₹)"]
    ws_orders.append(order_headers)
    ws_orders.row_dimensions[1].height = 24
    for col_num in range(1, len(order_headers) + 1):
        cell = ws_orders.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = order_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-fit columns
    for ws in [ws_catalog, ws_orders]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)
    
    ws_catalog.column_dimensions['A'].width = 12

    try:
        wb.save(target_path)
        print(f"Generated AADHAN FIRE WORKS workbook at {target_path}")
    except PermissionError:
        alt_path = target_path.replace(".xlsx", "_Updated.xlsx")
        wb.save(alt_path)
        print(f"File locked, saved updated workbook at {alt_path}")

    return items

if __name__ == "__main__":
    extract_and_generate()
